import importlib.util
import json
import tempfile
import unittest
import zipfile
from pathlib import Path


SCRIPT = Path(__file__).parents[1] / "scripts" / "generate_xlsm.py"
SPEC = importlib.util.spec_from_file_location("generate_xlsm", SCRIPT)
MODULE = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(MODULE)

VERIFY_SCRIPT = Path(__file__).parents[1] / "scripts" / "verify_xlsm.py"
VERIFY_SPEC = importlib.util.spec_from_file_location("verify_xlsm", VERIFY_SCRIPT)
VERIFY_MODULE = importlib.util.module_from_spec(VERIFY_SPEC)
VERIFY_SPEC.loader.exec_module(VERIFY_MODULE)


class MarkdownParsingTests(unittest.TestCase):
    def test_ignores_headings_inside_backtick_fences(self):
        source = "# API\nBody\n```bash\n# shell comment\necho ok\n```\n## Result\nDone\n"
        sections = self._parse(source)

        self.assertEqual(["API", "Result"], [item["heading"] for item in sections])
        self.assertIn("# shell comment", sections[0]["content"])

    def test_ignores_headings_inside_tilde_fences(self):
        source = "# API\n~~~~text\n## not a section\n~~~~\n## Result\nDone\n"
        sections = self._parse(source)

        self.assertEqual(["API", "Result"], [item["heading"] for item in sections])

    def _parse(self, content):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "source.md"
            path.write_text(content, encoding="utf-8")
            return MODULE.parse_markdown(path)


class HTMLParsingTests(unittest.TestCase):
    def test_extracts_section_id_heading_and_ignores_script(self):
        html = (
            '<section id="account" data-requirement-ids="REQ-001,REQ-002" '
            'data-source-evidence="src/account.py:create"><h2>Account</h2><p>Form body</p>'
            '<script>secret()</script></section>'
            '<section><h2>Result</h2><p>Done</p></section>'
        )
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "mock.html"
            path.write_text(html, encoding="utf-8")
            sections = MODULE.parse_html(path)

        self.assertEqual(["Account", "Result"], [item["heading"] for item in sections])
        self.assertEqual("#account", sections[0]["selector"])
        self.assertEqual("section:nth-of-type(2)", sections[1]["selector"])
        self.assertEqual("REQ-001,REQ-002", sections[0]["requirement_ids"])
        self.assertEqual("src/account.py:create", sections[0]["source_evidence"])
        self.assertNotIn("secret", sections[0]["content"])


class WorkbookSafetyTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise unittest.SkipTest("openpyxl is not installed") from exc
        cls.workbook_class = Workbook

    def test_rejects_existing_output_without_overwrite(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template = self._macro_template(root)
            output = root / "existing.xlsm"
            output.write_bytes(b"existing-result")

            with self.assertRaises(FileExistsError):
                MODULE.write_workbook(template, output, self._mapping(), self._rows())

            self.assertEqual(b"existing-result", output.read_bytes())

    def test_failure_does_not_replace_existing_output(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template = self._macro_template(root)
            output = root / "existing.xlsm"
            output.write_bytes(b"existing-result")
            mapping = self._mapping()
            mapping["sheet"] = "Missing"

            with self.assertRaises(ValueError):
                MODULE.write_workbook(template, output, mapping, self._rows(), overwrite=True)

            self.assertEqual(b"existing-result", output.read_bytes())

    def test_generates_workbook_and_preserves_vba_payload(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template = self._macro_template(root)
            output = root / "generated.xlsm"

            MODULE.write_workbook(template, output, self._mapping(), self._rows())

            self.assertEqual(MODULE.vba_hash(template), MODULE.vba_hash(output))
            workbook = load_workbook(output, keep_vba=True)
            sheet = workbook["Sections"]
            self.assertEqual("Heading", sheet["C2"].value)
            self.assertEqual("Body", sheet["D2"].value)
            self.assertEqual("=1+1", sheet["Z1"].value)
            self.assertEqual("0000FF00", sheet["Z1"].fill.fgColor.rgb)
            self.assertIn("X2:Y2", {str(item) for item in sheet.merged_cells.ranges})
            self.assertEqual("'Sections'!$Z$1", workbook.defined_names["TemplateAnchor"].attr_text)
            MODULE.close_workbook(workbook)

            errors = VERIFY_MODULE.verify(
                template, output, self._mapping(), expected_sections=1
            )
            self.assertEqual([], errors)

    def test_verifier_detects_change_outside_writable_range(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template = self._macro_template(root)
            output = root / "generated.xlsm"
            MODULE.write_workbook(template, output, self._mapping(), self._rows())

            workbook = load_workbook(output, keep_vba=True)
            workbook["Sections"]["Z1"] = "=9+9"
            workbook.save(output)
            MODULE.close_workbook(workbook)

            errors = VERIFY_MODULE.verify(
                template, output, self._mapping(), expected_sections=1
            )
            self.assertIn("Unexpected value change: Sections!Z1", errors)

    def test_verifier_detects_wrong_mapped_content(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template = self._macro_template(root)
            output = root / "generated.xlsm"
            rows = self._rows()
            MODULE.write_workbook(template, output, self._mapping(), rows)

            workbook = load_workbook(output, keep_vba=True)
            workbook["Sections"]["D2"] = "Wrong body"
            workbook.save(output)
            MODULE.close_workbook(workbook)

            errors = VERIFY_MODULE.verify(
                template,
                output,
                self._mapping(),
                expected_sections=1,
                expected_rows=rows,
            )
            self.assertTrue(any("Mapped value differs: Sections!D2" in error for error in errors))

    def test_embeds_screenshot_from_manifest(self):
        from openpyxl import load_workbook
        from openpyxl.utils.units import EMU_to_pixels
        from PIL import Image as PILImage

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template = self._macro_template(root)
            output = root / "generated.xlsm"
            screenshot = root / "section.png"
            PILImage.new("RGB", (32, 18), "white").save(screenshot)
            manifest = root / "screenshots.json"
            manifest.write_text(json.dumps({"sections": [{
                "source_file": "source.html",
                "selector": "#account",
                "render_mode": "image-and-text",
                "image": "section.png",
            }]}), encoding="utf-8")
            rows = self._rows()
            rows[0]["source_file"] = "source.html"
            rows[0]["selector"] = "#account"
            MODULE.attach_screenshots(rows, manifest)
            mapping = self._mapping()
            mapping["image"] = {
                "column": "H",
                "width": 320,
                "height": 180,
                "row_height_points": 140,
                "required": True,
            }

            MODULE.write_workbook(template, output, mapping, rows)

            workbook = load_workbook(output, keep_vba=True)
            sheet = workbook["Sections"]
            self.assertEqual(1, len(sheet._images))
            self.assertEqual(320, EMU_to_pixels(sheet._images[0].anchor.ext.cx))
            self.assertEqual(180, EMU_to_pixels(sheet._images[0].anchor.ext.cy))
            self.assertEqual(140, sheet.row_dimensions[2].height)
            MODULE.close_workbook(workbook)

            errors = VERIFY_MODULE.verify(
                template,
                output,
                mapping,
                expected_sections=1,
                manifest_path=manifest,
            )
            self.assertEqual([], errors)

    def test_mixes_text_and_html_rendering_image_sections(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            screenshot = root / "screen.png"
            screenshot.write_bytes(b"image-placeholder")
            manifest = root / "sections.json"
            manifest.write_text(json.dumps({"sections": [
                {
                    "source_file": "mock.html",
                    "selector": "#screen",
                    "render_mode": "image",
                    "image": "screen.png",
                },
                {
                    "source_file": "mock.html",
                    "selector": "#rules",
                    "render_mode": "text",
                },
            ]}), encoding="utf-8")
            rows = [
                {"source_file": "mock.html", "selector": "#screen"},
                {"source_file": "mock.html", "selector": "#rules"},
                {"source_file": "mock.html", "selector": "#default"},
            ]

            MODULE.attach_screenshots(rows, manifest)

            self.assertEqual("image", rows[0]["render_mode"])
            self.assertEqual(screenshot.resolve(), rows[0]["screenshot"])
            self.assertEqual("text", rows[1]["render_mode"])
            self.assertIsNone(rows[1]["screenshot"])
            self.assertEqual("text", rows[2]["render_mode"])

    def _macro_template(self, root):
        from openpyxl.styles import PatternFill
        from openpyxl.workbook.defined_name import DefinedName

        xlsx = root / "template.xlsx"
        xlsm = root / "template.xlsm"
        workbook = self.workbook_class()
        sheet = workbook.active
        sheet.title = "Sections"
        sheet["Z1"] = "=1+1"
        sheet["Z1"].fill = PatternFill(fill_type="solid", fgColor="00FF00")
        sheet.merge_cells("X2:Y2")
        workbook.defined_names.add(
            DefinedName("TemplateAnchor", attr_text="'Sections'!$Z$1")
        )
        workbook.save(xlsx)
        workbook.close()

        with zipfile.ZipFile(xlsx, "r") as source, zipfile.ZipFile(xlsm, "w") as target:
            for item in source.infolist():
                target.writestr(item, source.read(item.filename))
            target.writestr("xl/vbaProject.bin", b"test-vba-payload")
        return xlsm

    @staticmethod
    def _mapping():
        return {
            "sheet": "Sections",
            "start_row": 2,
            "columns": {"heading": "C", "content": "D"},
        }

    @staticmethod
    def _rows():
        return [{
            "level": 1,
            "heading": "Heading",
            "content": "Body",
            "source_file": "source.md",
            "source_location": "line 1",
        }]


if __name__ == "__main__":
    unittest.main()
