#!/usr/bin/env python3
import argparse
import importlib.util
import json
import sys
from pathlib import Path


def load_generator_module():
    script = Path(__file__).with_name("generate_xlsm.py")
    spec = importlib.util.spec_from_file_location("source_to_xlsm_generator", script)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def close_workbook(workbook):
    vba_archive = getattr(workbook, "vba_archive", None)
    if vba_archive is not None:
        vba_archive.close()
        workbook.vba_archive = None
    workbook.close()


def vba_hash(path):
    import hashlib
    import zipfile

    with zipfile.ZipFile(path) as archive:
        try:
            payload = archive.read("xl/vbaProject.bin")
        except KeyError as exc:
            raise ValueError(f"VBA project not found in {path}") from exc
    return hashlib.sha256(payload).hexdigest()


def defined_names(workbook):
    return {
        name: (
            item.attr_text,
            item.localSheetId,
            item.hidden,
            item.function,
            item.vbProcedure,
        )
        for name, item in workbook.defined_names.items()
    }


def expected_image_count(manifest_path):
    if manifest_path is None:
        return 0
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    sections = payload.get("sections", payload.get("screenshots", []))
    return sum(
        1
        for item in sections
        if item.get("render_mode", "image-and-text") in {"image", "image-and-text"}
    )


def verify(template_path, output_path, mapping, expected_sections, manifest_path=None, expected_rows=None):
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import column_index_from_string
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to verify .xlsm files") from exc

    errors = []
    template = load_workbook(template_path, keep_vba=True, keep_links=True)
    output = load_workbook(output_path, keep_vba=True, keep_links=True)
    try:
        if template.sheetnames != output.sheetnames:
            errors.append(f"Sheet order differs: {template.sheetnames} != {output.sheetnames}")
        if defined_names(template) != defined_names(output):
            errors.append("Workbook defined names differ")
        if vba_hash(template_path) != vba_hash(output_path):
            errors.append("VBA project differs")

        sheet_name = mapping["sheet"]
        if sheet_name not in template.sheetnames or sheet_name not in output.sheetnames:
            errors.append(f"Mapped worksheet is missing: {sheet_name}")
            return errors

        start_row = int(mapping["start_row"])
        end_row = start_row + expected_sections - 1
        if mapping.get("clear_existing"):
            end_row = max(end_row, int(mapping.get("clear_to_row", end_row)))
        writable_columns = {
            column_index_from_string(column)
            for column in mapping["columns"].values()
        }

        for name in template.sheetnames:
            template_sheet = template[name]
            output_sheet = output[name]
            template_merges = {str(item) for item in template_sheet.merged_cells.ranges}
            output_merges = {str(item) for item in output_sheet.merged_cells.ranges}
            if template_merges != output_merges:
                errors.append(f"Merged cells differ in {name}")

            max_row = max(template_sheet.max_row, output_sheet.max_row)
            max_column = max(template_sheet.max_column, output_sheet.max_column)
            for row in range(1, max_row + 1):
                for column in range(1, max_column + 1):
                    writable = (
                        name == sheet_name
                        and start_row <= row <= end_row
                        and column in writable_columns
                    )
                    if writable:
                        continue
                    before = template_sheet.cell(row, column)
                    after = output_sheet.cell(row, column)
                    if before.value != after.value:
                        errors.append(f"Unexpected value change: {name}!{after.coordinate}")
                    if before._style != after._style:
                        errors.append(f"Unexpected style change: {name}!{after.coordinate}")

        output_sheet = output[sheet_name]
        heading_column = mapping["columns"].get("heading")
        if heading_column:
            populated = sum(
                1
                for row in range(start_row, start_row + expected_sections)
                if output_sheet[f"{heading_column}{row}"].value not in {None, ""}
            )
            if populated != expected_sections:
                errors.append(
                    f"Section count differs: expected {expected_sections}, found {populated}"
                )

        if expected_rows is not None:
            generator = load_generator_module()
            for index, section in enumerate(expected_rows, 1):
                row_number = start_row + index - 1
                values = generator.section_values(section, index)
                for field, column in mapping["columns"].items():
                    expected = values[field]
                    actual = output_sheet[f"{column}{row_number}"].value
                    blank_equivalent = expected == "" and actual is None
                    if actual != expected and not blank_equivalent:
                        errors.append(
                            f"Mapped value differs: {sheet_name}!{column}{row_number} "
                            f"({field}) expected {expected!r}, found {actual!r}"
                        )

        expected_images = expected_image_count(manifest_path)
        template_images = len(template[sheet_name]._images)
        output_images = len(output_sheet._images)
        if output_images != template_images + expected_images:
            errors.append(
                f"Image count differs: expected {template_images + expected_images}, found {output_images}"
            )
    finally:
        close_workbook(template)
        close_workbook(output)
    return errors


def main():
    parser = argparse.ArgumentParser(description="Verify a generated .xlsm against its template contract.")
    parser.add_argument("--template", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--mapping", required=True, type=Path)
    parser.add_argument("--expected-sections", required=True, type=int)
    parser.add_argument("--manifest", type=Path)
    parser.add_argument("--input", required=True, action="append", type=Path)
    args = parser.parse_args()

    required = [args.template, args.output, args.mapping, *args.input]
    if args.manifest:
        required.append(args.manifest)
    for path in required:
        if not path.is_file():
            raise FileNotFoundError(path)
    if args.template.suffix.lower() != ".xlsm" or args.output.suffix.lower() != ".xlsm":
        raise ValueError("Template and output must use the .xlsm extension")
    if args.expected_sections < 1:
        raise ValueError("--expected-sections must be positive")

    mapping = json.loads(args.mapping.read_text(encoding="utf-8"))
    generator = load_generator_module()
    expected_rows = generator.parse_inputs(args.input)
    generator.attach_screenshots(expected_rows, args.manifest)
    if len(expected_rows) != args.expected_sections:
        raise ValueError(
            f"--expected-sections differs from parsed input: {args.expected_sections} != {len(expected_rows)}"
        )
    errors = verify(
        args.template,
        args.output,
        mapping,
        args.expected_sections,
        args.manifest,
        expected_rows,
    )
    result = {
        "result": "PASS" if not errors else "FAIL",
        "template": str(args.template),
        "output": str(args.output),
        "expected_sections": args.expected_sections,
        "errors": errors,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if not errors else 1


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
