#!/usr/bin/env python3
import argparse
import hashlib
import io
import json
import re
import shutil
import sys
import tempfile
import zipfile
from html.parser import HTMLParser
from pathlib import Path


class SectionHTMLParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.sections = []
        self.current = None
        self.heading_level = None
        self.heading_parts = []
        self.content_parts = []
        self.section_depth = 0
        self.section_count = 0
        self.skip_depth = 0

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag in {"script", "style", "template"}:
            self.skip_depth += 1
            return
        if self.skip_depth:
            return
        if tag == "section":
            if self.section_depth == 0:
                self._finish_section()
                self.section_count += 1
                attributes = dict(attrs)
                section_id = attributes.get("id")
                selector = f'#{section_id}' if section_id else f"section:nth-of-type({self.section_count})"
                self.current = {
                    "level": 1,
                    "heading": "",
                    "selector": selector,
                    "requirement_ids": attributes.get(
                        "data-requirement-ids", attributes.get("data-req-id", "")
                    ),
                    "source_evidence": attributes.get("data-source-evidence", ""),
                }
                self.content_parts = []
            self.section_depth += 1
            return
        match = re.fullmatch(r"h([1-6])", tag.lower())
        if match:
            if self.section_depth == 0:
                self._finish_section()
            self.heading_level = int(match.group(1))
            self.heading_parts = []
            if self.section_depth == 0:
                self.content_parts = []

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag in {"script", "style", "template"}:
            if self.skip_depth:
                self.skip_depth -= 1
            return
        if self.skip_depth:
            return
        if self.heading_level and tag.lower() == f"h{self.heading_level}":
            heading = " ".join(self.heading_parts).strip()
            if self.current is None:
                self.current = {
                    "level": self.heading_level,
                    "heading": heading,
                    "selector": f"h{self.heading_level}:nth-of-type({len(self.sections) + 1})",
                }
            elif not self.current["heading"]:
                self.current["level"] = self.heading_level
                self.current["heading"] = heading
            self.heading_level = None
            return
        if tag == "section" and self.section_depth:
            self.section_depth -= 1
            if self.section_depth == 0:
                self._finish_section()

    def handle_data(self, data):
        if self.skip_depth:
            return
        value = " ".join(data.split())
        if not value:
            return
        if self.heading_level:
            self.heading_parts.append(value)
        elif self.current:
            self.content_parts.append(value)

    def close(self):
        super().close()
        self._finish_section()

    def _finish_section(self):
        if self.current:
            if not self.current["heading"]:
                self.current["heading"] = f"Section {len(self.sections) + 1}"
            self.current["content"] = "\n".join(self.content_parts).strip()
            self.sections.append(self.current)
        self.current = None
        self.content_parts = []


def sha256(path):
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def vba_hash(path):
    with zipfile.ZipFile(path) as archive:
        try:
            data = archive.read("xl/vbaProject.bin")
        except KeyError as exc:
            raise ValueError(f"VBA project not found in {path}") from exc
    return hashlib.sha256(data).hexdigest()


def parse_markdown(path):
    sections = []
    current = None
    heading_pattern = re.compile(r"^(#{1,6})\s+(.+?)\s*$")
    fence_pattern = re.compile(r"^ {0,3}(`{3,}|~{3,})")
    active_fence = None
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        fence_match = fence_pattern.match(line)
        if fence_match:
            marker = fence_match.group(1)
            if active_fence is None:
                active_fence = (marker[0], len(marker))
            elif marker[0] == active_fence[0] and len(marker) >= active_fence[1]:
                active_fence = None
            if current:
                current["lines"].append(line)
            continue

        if active_fence is not None:
            if current:
                current["lines"].append(line)
            continue

        match = heading_pattern.match(line)
        if match:
            if current:
                current["content"] = "\n".join(current.pop("lines")).strip()
                sections.append(current)
            current = {
                "level": len(match.group(1)),
                "heading": match.group(2).strip(),
                "source_location": f"line {line_number}",
                "lines": [],
            }
        elif current:
            current["lines"].append(line)
    if current:
        current["content"] = "\n".join(current.pop("lines")).strip()
        sections.append(current)
    return sections


def parse_html(path):
    parser = SectionHTMLParser()
    parser.feed(path.read_text(encoding="utf-8"))
    parser.close()
    for section in parser.sections:
        section["source_location"] = section["selector"]
    return parser.sections


def parse_inputs(paths):
    rows = []
    for path in paths:
        suffix = path.suffix.lower()
        if suffix == ".md":
            sections = parse_markdown(path)
        elif suffix in {".html", ".htm"}:
            sections = parse_html(path)
        else:
            raise ValueError(f"Unsupported input format: {path}")
        for section in sections:
            section["source_file"] = str(path)
            rows.append(section)
    return rows


def attach_screenshots(rows, manifest_path):
    if manifest_path is None:
        for row in rows:
            row["render_mode"] = "text"
        return
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    definitions = {}
    for item in payload.get("sections", payload.get("screenshots", [])):
        key = (str(Path(item["source_file"])), item["selector"])
        mode = item.get("render_mode", "image-and-text")
        if mode not in {"text", "image", "image-and-text"}:
            raise ValueError(f"Unsupported render_mode: {mode}")
        image_path = None
        if item.get("image"):
            image_path = Path(item["image"])
            if not image_path.is_absolute():
                image_path = manifest_path.parent / image_path
            image_path = image_path.resolve()
        definitions[key] = {
            "render_mode": mode,
            "screenshot": image_path,
            "requirement_ids": item.get("requirement_ids"),
            "source_evidence": item.get("source_evidence"),
        }
    for row in rows:
        definition = definitions.get(
            (row["source_file"], row.get("selector")),
            {"render_mode": "text", "screenshot": None},
        )
        row["render_mode"] = definition["render_mode"]
        row["screenshot"] = definition["screenshot"]
        for field in ("requirement_ids", "source_evidence"):
            if definition.get(field) is not None:
                row[field] = definition[field]


def copy_row_style(sheet, source_row, target_row, columns):
    from copy import copy

    for column in columns:
        source = sheet[f"{column}{source_row}"]
        target = sheet[f"{column}{target_row}"]
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def close_workbook(workbook):
    vba_archive = getattr(workbook, "vba_archive", None)
    if vba_archive is not None:
        vba_archive.close()
        workbook.vba_archive = None
    workbook.close()


def section_values(section, index):
    render_mode = section.get("render_mode", "text")
    return {
        "index": index,
        "level": section["level"],
        "heading": section["heading"],
        "content": "" if render_mode == "image" else section["content"],
        "source_file": section["source_file"],
        "source_location": section["source_location"],
        "selector": section.get("selector", ""),
        "render_mode": render_mode,
        "requirement_ids": section.get("requirement_ids", ""),
        "source_evidence": section.get("source_evidence", ""),
    }


def write_workbook(template, output, mapping, rows, overwrite=False):
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to generate .xlsm files") from exc

    output.parent.mkdir(parents=True, exist_ok=True)
    if output.exists() and not overwrite:
        raise FileExistsError(f"Output already exists; pass --overwrite to replace it: {output}")

    temporary = None
    workbook = None
    try:
        with tempfile.NamedTemporaryFile(
            dir=output.parent, prefix=f".{output.stem}-", suffix=".xlsm", delete=False
        ) as stream:
            temporary = Path(stream.name)
        shutil.copy2(template, temporary)
        workbook = load_workbook(temporary, keep_vba=True, keep_links=True)
        sheet_name = mapping["sheet"]
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Worksheet not found: {sheet_name}")
        sheet = workbook[sheet_name]
        columns = mapping["columns"]
        start_row = int(mapping["start_row"])

        if mapping.get("clear_existing"):
            clear_to = int(mapping.get("clear_to_row", sheet.max_row))
            for row_number in range(start_row, clear_to + 1):
                for column in columns.values():
                    sheet[f"{column}{row_number}"].value = None

        style_row = mapping.get("copy_style_from_row")
        for index, section in enumerate(rows, 1):
            row_number = start_row + index - 1
            if style_row and row_number != int(style_row):
                copy_row_style(sheet, int(style_row), row_number, columns.values())
            render_mode = section.get("render_mode", "text")
            if render_mode not in {"text", "image", "image-and-text"}:
                raise ValueError(f"Unsupported render_mode: {render_mode}")
            values = section_values(section, index)
            for field, column in columns.items():
                if field not in values:
                    raise ValueError(f"Unsupported column field: {field}")
                sheet[f"{column}{row_number}"] = values[field]

            image_mapping = mapping.get("image")
            screenshot = section.get("screenshot")
            image_required = render_mode in {"image", "image-and-text"}
            if image_required and image_mapping is None:
                raise ValueError("Image mapping is required for image render modes")
            if image_mapping:
                if screenshot is None and image_required:
                    raise ValueError(f"HTML rendering image not found: {section['source_file']}::{section.get('selector', '')}")
                if screenshot is not None:
                    if not screenshot.is_file():
                        raise FileNotFoundError(screenshot)
                    image = Image(io.BytesIO(screenshot.read_bytes()))
                    if image_mapping.get("width"):
                        image.width = int(image_mapping["width"])
                    if image_mapping.get("height"):
                        image.height = int(image_mapping["height"])
                    sheet.add_image(image, f"{image_mapping['column']}{row_number}")
                    if image_mapping.get("row_height_points"):
                        sheet.row_dimensions[row_number].height = float(image_mapping["row_height_points"])

        workbook.save(temporary)
        close_workbook(workbook)
        workbook = None
        if vba_hash(temporary) != vba_hash(template):
            raise RuntimeError("VBA project differs between template and generated workbook")
        temporary.replace(output)
        temporary = None
    finally:
        if workbook is not None:
            close_workbook(workbook)
        if temporary is not None:
            temporary.unlink(missing_ok=True)


def main():
    parser = argparse.ArgumentParser(description="Populate a copied .xlsm template from Markdown or HTML sections.")
    parser.add_argument("--template", required=True, type=Path)
    parser.add_argument("--mapping", required=True, type=Path)
    parser.add_argument("--input", required=True, action="append", type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--screenshots", type=Path)
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    if args.template.suffix.lower() != ".xlsm" or args.output.suffix.lower() != ".xlsm":
        raise ValueError("Template and output must use the .xlsm extension")
    if args.template.resolve() == args.output.resolve():
        raise ValueError("Output must not overwrite the template")
    required_paths = [args.template, args.mapping, *args.input]
    if args.screenshots:
        required_paths.append(args.screenshots)
    for path in required_paths:
        if not path.is_file():
            raise FileNotFoundError(path)

    template_file_hash = sha256(args.template)
    template_vba_hash = vba_hash(args.template)
    mapping = json.loads(args.mapping.read_text(encoding="utf-8"))
    rows = parse_inputs(args.input)
    attach_screenshots(rows, args.screenshots)
    if not rows:
        raise ValueError("No sections found in the input files")

    write_workbook(args.template, args.output, mapping, rows, overwrite=args.overwrite)

    if sha256(args.template) != template_file_hash:
        raise RuntimeError("Template was modified during generation")
    if vba_hash(args.output) != template_vba_hash:
        raise RuntimeError("VBA project differs between template and output")

    print(json.dumps({
        "output": str(args.output),
        "sections": len(rows),
        "template_sha256": template_file_hash,
        "vba_sha256": template_vba_hash,
        "sources": [str(path) for path in args.input],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
